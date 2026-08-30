import os
import csv
import io
import re
import secrets
from functools import wraps
from datetime import date, datetime

from dotenv import load_dotenv
load_dotenv()

import psycopg2
import psycopg2.extras
from flask import Flask, request, jsonify, session, redirect, url_for, render_template, send_from_directory
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash, check_password_hash

import certificates as certs

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, "templates"), static_url_path="")

limiter = Limiter(key_func=get_remote_address)
limiter.init_app(app)

SECRET_KEY = os.environ.get("SECRET_KEY")
if not SECRET_KEY:
    SECRET_KEY = secrets.token_hex(32)
    print("WARNING: SECRET_KEY not set; using a random value (sessions reset on every restart)")
app.secret_key = SECRET_KEY

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")
if not ADMIN_PASSWORD:
    ADMIN_PASSWORD = secrets.token_hex(16)
    print("WARNING: ADMIN_PASSWORD not set; admin login is disabled until you set it")

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    print("WARNING: DATABASE_URL not set; registrations and admin features are disabled until you configure a Postgres connection.")

DB_AVAILABLE = bool(DATABASE_URL)

BOOTCAMP_DAYS = [
    {"number": 1, "date": "2026-08-17", "label": "Day 1 - Mon, Aug 17"},
    {"number": 2, "date": "2026-08-18", "label": "Day 2 - Tue, Aug 18"},
    {"number": 3, "date": "2026-08-19", "label": "Day 3 - Wed, Aug 19"},
    {"number": 4, "date": "2026-08-20", "label": "Day 4 - Thu, Aug 20"},
    {"number": 5, "date": "2026-08-21", "label": "Day 5 - Fri, Aug 21"},
    {"number": 6, "date": "2026-08-22", "label": "Day 6 - Sat, Aug 22"},
    {"number": 7, "date": "2026-08-23", "label": "Day 7 - Sun, Aug 23"},
    {"number": 8, "date": "2026-08-24", "label": "Day 8 - Mon, Aug 24"},
]


def get_db():
    if not DB_AVAILABLE:
        raise RuntimeError("DATABASE_URL is not configured.")
    return psycopg2.connect(DATABASE_URL, sslmode="require")


def ensure_db():
    if not DB_AVAILABLE:
        return False
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS registrations (
                    id SERIAL PRIMARY KEY,
                    fullname TEXT NOT NULL,
                    email TEXT NOT NULL,
                    phone TEXT NOT NULL,
                    level TEXT NOT NULL,
                    department TEXT NOT NULL,
                    expectation TEXT NOT NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT (NOW() AT TIME ZONE 'Africa/Lagos')
                )
            """)
            cur.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_registrations_email_lower "
                "ON registrations (LOWER(email))"
            )
            cur.execute("""
                CREATE TABLE IF NOT EXISTS attendance (
                    id SERIAL PRIMARY KEY,
                    registration_id INTEGER NOT NULL REFERENCES registrations(id) ON DELETE CASCADE,
                    day_number INTEGER NOT NULL,
                    attendance_date DATE NOT NULL,
                    marked_at TIMESTAMP NOT NULL DEFAULT (NOW() AT TIME ZONE 'Africa/Lagos'),
                    UNIQUE(registration_id, day_number)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS admins (
                    id SERIAL PRIMARY KEY,
                    username TEXT NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'admin',
                    created_at TIMESTAMP NOT NULL DEFAULT (NOW() AT TIME ZONE 'Africa/Lagos'),
                    UNIQUE(username)
                )
            """)
            cur.execute("SELECT COUNT(*) FROM admins")
            if cur.fetchone()[0] == 0:
                cur.execute(
                    "INSERT INTO admins (username, password_hash, role) VALUES (%s, %s, %s)",
                    ("admin", generate_password_hash(ADMIN_PASSWORD), "super_admin")
                )
            cur.execute("""
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS certificates (
                    id SERIAL PRIMARY KEY,
                    registration_id INTEGER NOT NULL UNIQUE REFERENCES registrations(id) ON DELETE CASCADE,
                    days_attended INTEGER NOT NULL DEFAULT 0,
                    eligible BOOLEAN NOT NULL DEFAULT FALSE,
                    certificate_number TEXT UNIQUE,
                    certificate_pdf BYTEA,
                    generated_at TIMESTAMP,
                    sent_at TIMESTAMP,
                    created_at TIMESTAMP NOT NULL DEFAULT (NOW() AT TIME ZONE 'Africa/Lagos')
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cert_template (
                    id INTEGER PRIMARY KEY,
                    image BYTEA NOT NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT (NOW() AT TIME ZONE 'Africa/Lagos')
                )
            """)
        conn.commit()
    return True


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not DB_AVAILABLE:
            return render_template(
                "admin_login.html",
                error="DATABASE_URL is not set. Configure a Postgres connection to enable registrations and admin access."
            ), 503
        if not session.get("admin_id"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


def super_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not DB_AVAILABLE:
            return jsonify({"error": "Database not available"}), 503
        if not session.get("admin_id"):
            return redirect(url_for("admin_login"))
        if session.get("admin_role") != "super_admin":
            return jsonify({"error": "Super admin access required"}), 403
        return f(*args, **kwargs)
    return decorated


@app.context_processor
def inject_admin_role():
    return dict(admin_role=session.get("admin_role", ""))


@app.route("/")
def home():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/style.css")
def serve_css():
    return send_from_directory(BASE_DIR, "style.css", mimetype="text/css")


@app.route("/script.js")
def serve_js():
    return send_from_directory(BASE_DIR, "script.js", mimetype="application/javascript")


@app.route("/images/<path:filename>")
def serve_image(filename):
    return send_from_directory(os.path.join(BASE_DIR, "images"), filename)


@app.route("/api/register", methods=["POST"])
@limiter.limit("5 per minute")
def register():
    if not DB_AVAILABLE:
        return jsonify({"error": "DATABASE_URL is not set. Registration is disabled until you configure a Postgres connection."}), 503
    data = request.get_json(force=True)
    required = ["fullname", "email", "phone", "level", "department", "expectation"]
    if not all(data.get(f, "").strip() for f in required):
        return jsonify({"error": "Missing fields"}), 400

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO registrations (fullname, email, phone, level, department, expectation) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    (data["fullname"].strip(), data["email"].strip().lower(), data["phone"].strip(),
                     data["level"], data["department"].strip(), data["expectation"])
                )
            conn.commit()
    except psycopg2.errors.UniqueViolation:
        return jsonify({"error": "This email is already registered."}), 409
    except psycopg2.Error:
        return jsonify({"error": "Could not save your registration. Please try again."}), 500

    return jsonify({"message": "Registration successful"}), 201


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if not username or not password:
            error = "Please enter both username and password."
        else:
            try:
                with get_db() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT id, username, password_hash, role FROM admins WHERE username = %s",
                            (username,)
                        )
                        admin = cur.fetchone()
                if admin and check_password_hash(admin[2], password):
                    session["admin_id"] = admin[0]
                    session["admin_username"] = admin[1]
                    session["admin_role"] = admin[3]
                    return redirect(url_for("admin_dashboard"))
                error = "Invalid username or password."
            except Exception:
                error = "Login failed. Please try again."
    return render_template("admin_login.html", error=error)


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_id", None)
    session.pop("admin_username", None)
    session.pop("admin_role", None)
    return redirect(url_for("admin_login"))


PER_PAGE = 25


def build_query(filters):
    search = filters.get("search")
    level_filter = filters.get("level")
    dept_filter = filters.get("department")

    query = "SELECT id, fullname, email, phone, level, department, expectation, created_at FROM registrations WHERE 1=1"
    params = []

    if search:
        query += " AND (fullname ILIKE %s OR email ILIKE %s OR department ILIKE %s OR phone ILIKE %s)"
        like = f"%{search}%"
        params.extend([like, like, like, like])

    if level_filter:
        query += " AND level = %s"
        params.append(level_filter)

    if dept_filter:
        query += " AND department = %s"
        params.append(dept_filter)

    return query, params


@app.route("/admin")
@admin_required
def admin_dashboard():
    search = request.args.get("search", "").strip()
    level_filter = request.args.get("level", "").strip()
    dept_filter = request.args.get("department", "").strip()

    filters = {"search": search, "level": level_filter, "department": dept_filter}
    where_query, params = build_query(filters)

    try:
        page = max(int(request.args.get("page", 1)), 1)
    except (TypeError, ValueError):
        page = 1

    with get_db() as conn:
        with conn.cursor() as cur:
            count_query = where_query.replace(
                "SELECT id, fullname, email, phone, level, department, expectation, created_at",
                "SELECT COUNT(*)"
            )
            cur.execute(count_query, params)
            filtered_count = cur.fetchone()[0]
            total_pages = max((filtered_count + PER_PAGE - 1) // PER_PAGE, 1)
            page = min(page, total_pages)

            cur.execute(where_query + " ORDER BY id DESC LIMIT %s OFFSET %s", params + [PER_PAGE, (page - 1) * PER_PAGE])
            rows = cur.fetchall()

            cur.execute("SELECT COUNT(*) FROM registrations")
            total = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM registrations WHERE created_at::date = (CURRENT_TIMESTAMP AT TIME ZONE 'Africa/Lagos')::date")
            today_count = cur.fetchone()[0]

            cur.execute("SELECT COUNT(DISTINCT department) FROM registrations WHERE department != ''")
            unique_depts = cur.fetchone()[0]

            cur.execute("SELECT DISTINCT department FROM registrations WHERE department != '' ORDER BY department")
            departments = [r[0] for r in cur.fetchall()]

    window = 2
    start = max(page - window, 1)
    end = min(page + window, total_pages)
    page_range = list(range(start, end + 1))

    return render_template("admin_dashboard.html",
        rows=rows, total=total, today_count=today_count,
        unique_depts=unique_depts, departments=departments,
        search=search, level_filter=level_filter, dept_filter=dept_filter,
        page=page, total_pages=total_pages, page_range=page_range,
        has_prev=page > 1, has_next=page < total_pages, per_page=PER_PAGE)


@app.route("/admin/export")
@admin_required
def export_csv():
    search = request.args.get("search", "").strip()
    level_filter = request.args.get("level", "").strip()
    dept_filter = request.args.get("department", "").strip()

    filters = {"search": search, "level": level_filter, "department": dept_filter}
    where_query, params = build_query(filters)
    query = "SELECT fullname, email, phone, level, department, expectation, created_at FROM registrations WHERE 1=1"
    if search or level_filter or dept_filter:
        query = where_query.replace(
            "SELECT id, fullname, email, phone, level, department, expectation, created_at",
            "SELECT fullname, email, phone, level, department, expectation, created_at"
        )
    query += " ORDER BY id DESC"

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(query, params)
            rows = cur.fetchall()

    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow(["Full Name", "Email", "Phone", "Level", "Department", "Expectation", "Timestamp"])
    for r in rows:
        writer.writerow(list(r))

    return app.response_class(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=registrations.csv"}
    )


@app.route("/admin/attendance")
@admin_required
def admin_attendance():
    try:
        selected_day = int(request.args.get("day", 4))
    except (TypeError, ValueError):
        selected_day = 4
    selected_day = max(1, min(selected_day, 8))

    day_info = next(d for d in BOOTCAMP_DAYS if d["number"] == selected_day)

    search = request.args.get("search", "").strip()

    with get_db() as conn:
        with conn.cursor() as cur:
            query = """
                SELECT r.id, r.fullname, r.email, r.phone, r.level, r.department,
                       a.id IS NOT NULL AS attended
                FROM registrations r
                LEFT JOIN attendance a ON a.registration_id = r.id AND a.day_number = %s
                WHERE 1=1
            """
            params = [selected_day]

            if search:
                query += " AND (r.fullname ILIKE %s OR r.email ILIKE %s OR r.phone ILIKE %s)"
                like = f"%{search}%"
                params.extend([like, like, like])

            query += " ORDER BY r.fullname ASC"
            cur.execute(query, params)
            rows = cur.fetchall()

            cur.execute("SELECT COUNT(*) FROM registrations")
            total_registered = cur.fetchone()[0]

            cur.execute(
                "SELECT COUNT(*) FROM attendance WHERE day_number = %s",
                (selected_day,)
            )
            attended_count = cur.fetchone()[0]

            cur.execute("""
                SELECT day_number, COUNT(*) FROM attendance
                GROUP BY day_number ORDER BY day_number
            """)
            day_counts = {r[0]: r[1] for r in cur.fetchall()}

    return render_template("admin_attendance.html",
        rows=rows, day_info=day_info, days=BOOTCAMP_DAYS,
        selected_day=selected_day, search=search,
        total_registered=total_registered, attended_count=attended_count,
        day_counts=day_counts)


@app.route("/admin/attendance/mark", methods=["POST"])
@admin_required
def mark_attendance():
    data = request.get_json(force=True)
    reg_id = data.get("registration_id")
    day = data.get("day_number")

    if not reg_id or not day:
        return jsonify({"error": "Missing registration_id or day_number"}), 400

    try:
        day = int(day)
        reg_id = int(reg_id)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid parameters"}), 400

    if day < 1 or day > 8:
        return jsonify({"error": "Day must be 1-8"}), 400

    day_info = next(d for d in BOOTCAMP_DAYS if d["number"] == day)

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO attendance (registration_id, day_number, attendance_date) "
                    "VALUES (%s, %s, %s) ON CONFLICT (registration_id, day_number) DO NOTHING",
                    (reg_id, day, day_info["date"])
                )
            conn.commit()
    except psycopg2.Error:
        return jsonify({"error": "Failed to mark attendance"}), 500

    return jsonify({"message": "Attendance marked"}), 201


@app.route("/admin/attendance/unmark", methods=["POST"])
@admin_required
def unmark_attendance():
    data = request.get_json(force=True)
    reg_id = data.get("registration_id")
    day = data.get("day_number")

    if not reg_id or not day:
        return jsonify({"error": "Missing registration_id or day_number"}), 400

    try:
        day = int(day)
        reg_id = int(reg_id)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid parameters"}), 400

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM attendance WHERE registration_id = %s AND day_number = %s",
                    (reg_id, day)
                )
            conn.commit()
    except psycopg2.Error:
        return jsonify({"error": "Failed to unmark attendance"}), 500

    return jsonify({"message": "Attendance removed"}), 200


@app.route("/admin/attendance/export/<int:day>")
@admin_required
def export_attendance(day):
    if day < 1 or day > 8:
        return "Invalid day", 400

    day_info = next(d for d in BOOTCAMP_DAYS if d["number"] == day)

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT r.fullname, r.email, r.phone, r.level, r.department
                FROM registrations r
                INNER JOIN attendance a ON a.registration_id = r.id AND a.day_number = %s
                ORDER BY r.fullname ASC
            """, (day,))
            rows = cur.fetchall()

    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow([f"Attendance - {day_info['label']}"])
    writer.writerow([])
    writer.writerow(["Full Name", "Email", "Phone", "Level", "Department"])
    for r in rows:
        writer.writerow(list(r))

    filename = f"attendance_day_{day}.csv"
    return app.response_class(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/admin/import/registrations", methods=["POST"])
@admin_required
def import_registrations():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only Excel files (.xlsx) are supported"}), 400

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

        name_idx = next((i for i, h in enumerate(headers) if h in ("name", "fullname", "full name")), None)
        phone_idx = next((i for i, h in enumerate(headers) if h in ("phone", "phone number", "phonenumber", "tel", "telephone")), None)
        dept_idx = next((i for i, h in enumerate(headers) if h in ("department", "dept", "faculty")), None)
        level_idx = next((i for i, h in enumerate(headers) if h in ("level", "level ", "class")), None)
        exp_idx = next((i for i, h in enumerate(headers) if h in ("expectation", "expect", "what do you expect")), None)
        email_idx = next((i for i, h in enumerate(headers) if h in ("email", "e-mail", "email address")), None)

        if name_idx is None or phone_idx is None:
            wb.close()
            return jsonify({"error": "Excel must have at least 'name' and 'phone' columns"}), 400

        inserted = 0
        skipped = 0
        errors = []

        with get_db() as conn:
            with conn.cursor() as cur:
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    name = str(row[name_idx]).strip() if row[name_idx] else ""
                    phone = str(row[phone_idx]).strip() if row[phone_idx] else ""
                    dept = str(row[dept_idx]).strip() if dept_idx is not None and row[dept_idx] else "Other"
                    level = str(row[level_idx]).strip() if level_idx is not None and row[level_idx] else "300"
                    expectation = str(row[exp_idx]).strip() if exp_idx is not None and row[exp_idx] else "Learn MATLAB basics"
                    email = str(row[email_idx]).strip().lower() if email_idx is not None and row[email_idx] else ""

                    if not name or not phone:
                        skipped += 1
                        continue

                    if not email:
                        email = f"import_{phone}@placeholder.local"

                    try:
                        cur.execute(
                            "INSERT INTO registrations (fullname, email, phone, level, department, expectation) "
                            "VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT (LOWER(email)) DO NOTHING",
                            (name, email, phone, level, dept, expectation)
                        )
                        if cur.rowcount > 0:
                            inserted += 1
                        else:
                            skipped += 1
                    except psycopg2.Error as e:
                        errors.append(f"Row {row_num}: {str(e)[:80]}")
                        skipped += 1

            conn.commit()
        wb.close()

        return jsonify({
            "message": f"Import complete: {inserted} registered, {skipped} skipped",
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors[:10]
        }), 200

    except Exception as e:
        return jsonify({"error": f"Failed to process file: {str(e)[:200]}"}), 500


@app.route("/admin/import/attendance", methods=["POST"])
@admin_required
def import_attendance():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    day = request.form.get("day_number")

    if not day:
        return jsonify({"error": "Please select a day"}), 400

    try:
        day = int(day)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid day number"}), 400

    if day < 1 or day > 8:
        return jsonify({"error": "Day must be 1-8"}), 400

    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only Excel files (.xlsx) are supported"}), 400

    day_info = next(d for d in BOOTCAMP_DAYS if d["number"] == day)

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

        name_idx = next((i for i, h in enumerate(headers) if h in ("name", "fullname", "full name")), None)
        phone_idx = next((i for i, h in enumerate(headers) if h in ("phone", "phone number", "phonenumber", "tel", "telephone")), None)
        email_idx = next((i for i, h in enumerate(headers) if h in ("email", "e-mail", "email address")), None)

        if name_idx is None:
            wb.close()
            return jsonify({"error": "Excel must have a 'name' column"}), 400

        marked = 0
        not_found = 0

        with get_db() as conn:
            with conn.cursor() as cur:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = str(row[name_idx]).strip() if row[name_idx] else ""
                    phone = str(row[phone_idx]).strip() if phone_idx is not None and row[phone_idx] else ""
                    email = str(row[email_idx]).strip().lower() if email_idx is not None and row[email_idx] else ""

                    if not name:
                        continue

                    reg_id = None

                    if email:
                        cur.execute("SELECT id FROM registrations WHERE LOWER(email) = %s", (email,))
                        result = cur.fetchone()
                        if result:
                            reg_id = result[0]

                    if not reg_id and phone:
                        cur.execute("SELECT id FROM registrations WHERE phone = %s", (phone,))
                        result = cur.fetchone()
                        if result:
                            reg_id = result[0]

                    if not reg_id:
                        cur.execute(
                            "SELECT id FROM registrations WHERE fullname ILIKE %s",
                            (name,)
                        )
                        result = cur.fetchone()
                        if result:
                            reg_id = result[0]

                    if reg_id:
                        cur.execute(
                            "INSERT INTO attendance (registration_id, day_number, attendance_date) "
                            "VALUES (%s, %s, %s) ON CONFLICT (registration_id, day_number) DO NOTHING",
                            (reg_id, day, day_info["date"])
                        )
                        if cur.rowcount > 0:
                            marked += 1
                    else:
                        not_found += 1

            conn.commit()
        wb.close()

        return jsonify({
            "message": f"Attendance import complete: {marked} marked present, {not_found} not found in registrations",
            "marked": marked,
            "not_found": not_found
        }), 200

    except Exception as e:
        return jsonify({"error": f"Failed to process file: {str(e)[:200]}"}), 500


@app.route("/admin/change-password", methods=["GET", "POST"])
@admin_required
def change_password():
    error = None
    success = None
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_pass = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")

        if not current or not new_pass or not confirm:
            error = "All fields are required."
        elif new_pass != confirm:
            error = "New passwords do not match."
        elif len(new_pass) < 6:
            error = "Password must be at least 6 characters."
        else:
            try:
                with get_db() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT id, password_hash FROM admins WHERE id = %s",
                            (session["admin_id"],)
                        )
                        admin = cur.fetchone()
                        if not admin or not check_password_hash(admin[1], current):
                            error = "Current password is incorrect."
                        else:
                            cur.execute(
                                "UPDATE admins SET password_hash = %s WHERE id = %s",
                                (generate_password_hash(new_pass), session["admin_id"])
                            )
                            conn.commit()
                            success = "Password changed successfully."
            except Exception:
                error = "Failed to change password. Please try again."
    return render_template("admin_change_password.html", error=error, success=success)


@app.route("/admin/manage-admins")
@super_admin_required
def manage_admins():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, username, role, created_at FROM admins ORDER BY id")
            admins = cur.fetchall()
    return render_template("admin_manage.html", admins=admins)


@app.route("/admin/create-admin", methods=["POST"])
@super_admin_required
def create_admin():
    data = request.get_json(force=True)
    username = data.get("username", "").strip()
    password = data.get("password", "")
    role = data.get("role", "admin")

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    if role not in ("admin", "super_admin"):
        return jsonify({"error": "Invalid role"}), 400

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO admins (username, password_hash, role) VALUES (%s, %s, %s)",
                    (username, generate_password_hash(password), role)
                )
            conn.commit()
    except psycopg2.errors.UniqueViolation:
        return jsonify({"error": "Username already exists"}), 409
    except psycopg2.Error:
        return jsonify({"error": "Failed to create admin"}), 500

    return jsonify({"message": f"Admin '{username}' created successfully"}), 201


@app.route("/admin/delete-admin", methods=["POST"])
@super_admin_required
def delete_admin():
    data = request.get_json(force=True)
    admin_id = data.get("admin_id")

    if not admin_id:
        return jsonify({"error": "Missing admin_id"}), 400

    try:
        admin_id = int(admin_id)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid admin_id"}), 400

    if admin_id == session.get("admin_id"):
        return jsonify({"error": "You cannot delete your own account"}), 400

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM admins WHERE id = %s", (admin_id,))
            conn.commit()
    except psycopg2.Error:
        return jsonify({"error": "Failed to delete admin"}), 500

    return jsonify({"message": "Admin deleted"}), 200


@app.route("/api/admin/register-user", methods=["POST"])
@admin_required
def admin_register_user():
    data = request.get_json(force=True)
    required = ["fullname", "email", "phone", "level", "department", "expectation"]
    if not all(data.get(f, "").strip() for f in required):
        return jsonify({"error": "All fields are required"}), 400

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO registrations (fullname, email, phone, level, department, expectation) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    (data["fullname"].strip(), data["email"].strip().lower(), data["phone"].strip(),
                     data["level"], data["department"].strip(), data["expectation"])
                )
            conn.commit()
    except psycopg2.errors.UniqueViolation:
        return jsonify({"error": "This email is already registered."}), 409
    except psycopg2.Error:
        return jsonify({"error": "Could not save registration."}), 500

    return jsonify({"message": "Registration successful"}), 201


@app.route("/api/admin/search-users")
@admin_required
def admin_search_users():
    q = request.args.get("q", "").strip()
    if not q or len(q) < 2:
        return jsonify([])

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, fullname, email, phone FROM registrations "
                    "WHERE fullname ILIKE %s OR email ILIKE %s OR phone ILIKE %s LIMIT 10",
                    (f"%{q}%", f"%{q}%", f"%{q}%")
                )
                results = [
                    {"id": r[0], "fullname": r[1], "email": r[2], "phone": r[3]}
                    for r in cur.fetchall()
                ]
    except Exception:
        results = []

    return jsonify(results)


@app.route("/admin/certificates")
@admin_required
def admin_certificates():
    with get_db() as conn:
        settings = certs.get_settings(conn)
        template = certs.get_template(conn) is not None
        participants = certs.eligible_participants(conn, int(settings["cert_min_days"]))

        sent_by_reg = {}
        with conn.cursor() as cur:
            cur.execute("SELECT registration_id, sent_at, generated_at, certificate_number FROM certificates")
            for reg_id, sent_at, generated_at, num in cur.fetchall():
                sent_by_reg[reg_id] = {"sent_at": sent_at, "generated_at": generated_at, "number": num}

    for p in participants:
        info = sent_by_reg.get(p["registration_id"])
        p["sent_at"] = info["sent_at"] if info else None
        p["generated_at"] = info["generated_at"] if info else None
        p["certificate_number"] = info["number"] if info else None

    eligible_count = sum(1 for p in participants if p["eligible"])
    generated_count = sum(1 for p in participants if p["generated_at"])
    sent_count = sum(1 for p in participants if p["sent_at"])

    return render_template("admin_certificates.html",
        settings=settings, has_template=template,
        participants=participants, days=BOOTCAMP_DAYS,
        eligible_count=eligible_count, generated_count=generated_count, sent_count=sent_count,
        smtp_ok=certs.smtp_configured())


@app.route("/admin/certificates/settings", methods=["POST"])
@admin_required
def certificates_settings():
    data = request.get_json(force=True)
    allowed = {
        "cert_min_days", "cert_name_x", "cert_name_y", "cert_name_size",
        "cert_name_color", "cert_font", "cert_event_name", "cert_theme",
        "cert_number_prefix", "cert_body"
    }
    with get_db() as conn:
        settings = certs.get_settings(conn)
        for k, v in data.items():
            if k in allowed and v is not None:
                settings[k] = str(v)
        try:
            creds = [not settings[k].strip() for k in ("cert_min_days", "cert_name_x", "cert_name_y", "cert_name_size", "cert_name_color")]
            if any(creds):
                return jsonify({"error": "Position, size, color and minimum days cannot be empty."}), 400
            if not (1 <= int(settings["cert_min_days"]) <= 8):
                return jsonify({"error": "Minimum days must be between 1 and 8."}), 400
        except ValueError:
            return jsonify({"error": "Numeric fields must be numbers."}), 400
        certs.save_settings(conn, settings)
    return jsonify({"message": "Certificate settings saved."}), 200


@app.route("/admin/certificates/template", methods=["POST"])
@admin_required
def certificates_upload_template():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename.lower().endswith((".png", ".jpg", ".jpeg")):
        return jsonify({"error": "Only PNG or JPG images are supported"}), 400
    data = file.read()
    if not data:
        return jsonify({"error": "Empty file"}), 400
    if len(data) > 15 * 1024 * 1024:
        return jsonify({"error": "Image too large (max 15 MB)"}), 400
    try:
        with get_db() as conn:
            certs.save_template(conn, data)
    except Exception:
        return jsonify({"error": "Failed to save template"}), 500
    return jsonify({"message": "Certificate template uploaded successfully."}), 200


@app.route("/admin/certificates/generate")
@admin_required
def certificates_generate_all():
    with get_db() as conn:
        settings = certs.get_settings(conn)
        template = certs.get_template(conn)
        if not template:
            return jsonify({"error": "No certificate template uploaded yet."}), 400

        participants = certs.eligible_participants(conn, int(settings["cert_min_days"]))
        eligible = [p for p in participants if p["eligible"]]

        with conn.cursor() as cur:
            cur.execute("SELECT registration_id, certificate_pdf FROM certificates")
            existing = {r[0]: r[1] for r in cur.fetchall()}

        generated = 0
        for p in eligible:
            reg_id = p["registration_id"]
            if existing.get(reg_id):
                generated += 1
                continue
            pdf = certs.render_certificate_pdf(template, p["fullname"], settings)
            number = certs.assign_next_certificate_number(conn, settings["cert_number_prefix"])
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO certificates (registration_id, days_attended, eligible, certificate_number, certificate_pdf, generated_at) "
                    "VALUES (%s, %s, %s, %s, %s, NOW() AT TIME ZONE 'Africa/Lagos') "
                    "ON CONFLICT (registration_id) DO UPDATE SET certificate_pdf = EXCLUDED.certificate_pdf, "
                    "certificate_number = COALESCE(certificates.certificate_number, EXCLUDED.certificate_number), "
                    "generated_at = NOW() AT TIME ZONE 'Africa/Lagos'",
                    (reg_id, p["days_attended"], True, number, psycopg2.Binary(pdf))
                )
            conn.commit()
            generated += 1

        not_eligible = sum(1 for p in participants if not p["eligible"])
    return jsonify({
        "message": f"Generated {generated} certificate{'s' if generated != 1 else ''}. {not_eligible} not eligible.",
        "generated": generated,
        "not_eligible": not_eligible
    }), 200


@app.route("/admin/certificates/generate/<int:reg_id>", methods=["POST"])
@admin_required
def certificates_generate_one(reg_id):
    with get_db() as conn:
        settings = certs.get_settings(conn)
        template = certs.get_template(conn)
        if not template:
            return jsonify({"error": "No certificate template uploaded yet."}), 400

        with conn.cursor() as cur:
            cur.execute(
                "SELECT fullname FROM registrations WHERE id = %s",
                (reg_id,)
            )
            row = cur.fetchone()
            if not row:
                return jsonify({"error": "Participant not found."}), 404
            fullname = certs.normalize_name(row[0])
            cur.execute(
                "SELECT COUNT(*) FROM attendance WHERE registration_id = %s",
                (reg_id,)
            )
            days = cur.fetchone()[0]

        pdf = certs.render_certificate_pdf(template, fullname, settings)
        number = certs.assign_next_certificate_number(conn, settings["cert_number_prefix"])

        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO certificates (registration_id, days_attended, eligible, certificate_number, certificate_pdf, generated_at) "
                "VALUES (%s, %s, %s, %s, %s, NOW() AT TIME ZONE 'Africa/Lagos') "
                "ON CONFLICT (registration_id) DO UPDATE SET certificate_pdf = EXCLUDED.certificate_pdf, "
                "certificate_number = COALESCE(certificates.certificate_number, EXCLUDED.certificate_number), "
                "generated_at = NOW() AT TIME ZONE 'Africa/Lagos'",
                (reg_id, days, days >= int(settings["cert_min_days"]), number, psycopg2.Binary(pdf))
            )
        conn.commit()
    return jsonify({"message": "Certificate generated.", "certificate_number": number}), 200


@app.route("/admin/certificates/send/<int:reg_id>", methods=["POST"])
@admin_required
def certificates_send_one(reg_id):
    if not certs.smtp_configured():
        return jsonify({"error": "SMTP is not configured. Set SMTP_HOST, SMTP_USER and SMTP_PASSWORD."}), 400

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT c.certificate_pdf, c.certificate_number, r.fullname, r.email "
                "FROM certificates c JOIN registrations r ON r.id = c.registration_id "
                "WHERE c.registration_id = %s",
                (reg_id,)
            )
            row = cur.fetchone()
            if not row or not row[0]:
                return jsonify({"error": "No generated certificate for this participant yet."}), 400
            pdf, number, fullname, email = row

        if certs.is_placeholder_email(email):
            return jsonify({"error": "Participant has no valid email address to send to."}), 400

        settings = certs.get_settings(conn)
        try:
            certs.send_certificate_email(email, certs.normalize_name(fullname), number, bytes(pdf), settings)
        except Exception as e:
            return jsonify({"error": f"Email failed: {str(e)[:160]}"}), 500

        with conn.cursor() as cur:
            cur.execute(
                "UPDATE certificates SET sent_at = NOW() AT TIME ZONE 'Africa/Lagos' WHERE registration_id = %s",
                (reg_id,)
            )
        conn.commit()
    return jsonify({"message": f"Certificate sent to {email}."}), 200


@app.route("/admin/certificates/download/<int:reg_id>")
@admin_required
def certificates_download(reg_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT c.certificate_pdf, r.fullname "
                "FROM certificates c JOIN registrations r ON r.id = c.registration_id "
                "WHERE c.registration_id = %s",
                (reg_id,)
            )
            row = cur.fetchone()
            if not row or not row[0]:
                return "Certificate not generated yet", 404
            pdf, fullname = row

    safe_name = re.sub(r"[^A-Za-z0-9 _-]", "", fullname).replace(" ", "_")
    return app.response_class(
        bytes(pdf),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=FUTMinna_MATLAB_Certificate_{safe_name}.pdf"}
    )


@app.route("/admin/certificates/preview")
@admin_required
def certificates_preview():
    with get_db() as conn:
        template = certs.get_template(conn)
        if not template:
            return "No template uploaded yet", 404
        settings = certs.get_settings(conn)
        sample = request.args.get("name", "John Doe")
        settings["cert_name_x"] = request.args.get("x", settings.get("cert_name_x"))
        settings["cert_name_y"] = request.args.get("y", settings.get("cert_name_y"))
        settings["cert_name_size"] = request.args.get("size", settings.get("cert_name_size"))
        settings["cert_name_color"] = request.args.get("color", settings.get("cert_name_color"))
        settings["cert_font"] = request.args.get("font", settings.get("cert_font"))
        png = certs.render_certificate_png(template, certs.normalize_name(sample), settings)
    return app.response_class(png, mimetype="image/png")


@app.route("/admin/certificates/export")
@admin_required
def certificates_export():
    with get_db() as conn:
        settings = certs.get_settings(conn)
        participants = certs.eligible_participants(conn, int(settings["cert_min_days"]))
        with conn.cursor() as cur:
            cur.execute("SELECT registration_id, certificate_number, sent_at, generated_at FROM certificates")
            info = {r[0]: r for r in cur.fetchall()}

    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow(["Full Name", "Email", "Days Attended", "Eligible", "Certificate Number", "Generated", "Sent"])
    for p in participants:
        i = info.get(p["registration_id"])
        writer.writerow([
            p["fullname"], p["email"], p["days_attended"], "Yes" if p["eligible"] else "No",
            i[1] if i else "", "Yes" if i and i[3] else "No", "Yes" if i and i[2] else "No"
        ])

    return app.response_class(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=certificates_eligibility.csv"}
    )


@app.route("/admin/certificates/eligible")
@admin_required
def certificates_eligible_json():
    with get_db() as conn:
        settings = certs.get_settings(conn)
        participants = certs.eligible_participants(conn, int(settings["cert_min_days"]))
        with conn.cursor() as cur:
            cur.execute("SELECT registration_id, sent_at, generated_at, certificate_number FROM certificates")
            info = {r[0]: r for r in cur.fetchall()}
        for p in participants:
            i = info.get(p["registration_id"])
            p["generated_at"] = i[2].isoformat() if i and i[2] else None
            p["sent_at"] = i[1].isoformat() if i and i[1] else None
            p["certificate_number"] = i[3] if i else None
            p["skippable"] = certs.is_placeholder_email(p["email"])
    return jsonify(participants)


if __name__ == "__main__":
    ensure_db()
    app.run(debug=True, port=5000)
